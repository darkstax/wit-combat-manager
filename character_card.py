"""Version-aware WIT character-card importers."""

from dataclasses import dataclass
import re
from typing import Any

import openpyxl

from models import RuleMode, Unit


@dataclass(frozen=True)
class CellSource:
    sheet: str
    cell: str


@dataclass(frozen=True)
class CharacterCardProfile:
    rule_mode: RuleMode
    required_sheets: frozenset[str]
    fields: dict[str, tuple[CellSource, ...]]
    defaults: dict[str, Any]


@dataclass(frozen=True)
class ImportResult:
    unit: Unit
    detected_rule_mode: RuleMode
    imported_fields: dict[str, Any]
    warnings: tuple[str, ...] = ()


V03_PROFILE = CharacterCardProfile(
    rule_mode=RuleMode.V0_3,
    required_sheets=frozenset({"角色卡", "简化卡", "光荣之路"}),
    fields={
        "name": (CellSource("角色卡", "H7"),),
        "max_hp": (CellSource("角色卡", "AH139"), CellSource("简化卡", "C5")),
        "physical_resist": (CellSource("角色卡", "AW139"), CellSource("简化卡", "I5")),
        "magic_resist": (CellSource("角色卡", "AZ139"), CellSource("简化卡", "K5")),
        "speed": (CellSource("角色卡", "AN139"), CellSource("简化卡", "M5")),
        "reaction_mobility": (CellSource("角色卡", "I21"), CellSource("简化卡", "G3")),
        "elemental_tenacity": (CellSource("角色卡", "AK139"), CellSource("简化卡", "E50")),
        "sp": (CellSource("角色卡", "BI139"), CellSource("简化卡", "M48")),
        "level": (CellSource("角色卡", "AX22"),),
        "profession": (CellSource("角色卡", "H14"),),
        "armor_type": (CellSource("角色卡", "AH141"),),
    },
    defaults={
        "speed": 0,
        "elemental_tenacity": 10,
        "sp": 0,
        "armor_type": "无甲",
    },
)

V12_PROFILE = CharacterCardProfile(
    rule_mode=RuleMode.V1_2,
    required_sheets=frozenset({"主卡", "战斗卡", "简化卡"}),
    fields={
        "name": (CellSource("主卡", "D3"),),
        "max_hp": (CellSource("简化卡", "H8"), CellSource("主卡", "AI24")),
        "physical_resist": (CellSource("简化卡", "K24"), CellSource("主卡", "AI25")),
        "magic_resist": (CellSource("简化卡", "K25"), CellSource("主卡", "AI26")),
        "speed": (CellSource("简化卡", "D3"), CellSource("主卡", "Q5")),
        "elemental_tenacity": (CellSource("简化卡", "P24"), CellSource("主卡", "AI29")),
        "weight": (CellSource("简化卡", "P25"), CellSource("主卡", "AI31")),
        "current_sp": (CellSource("简化卡", "N23"), CellSource("主卡", "AJ27")),
        "max_sp": (CellSource("简化卡", "N22"), CellSource("主卡", "AM27")),
        "current_stamina": (CellSource("简化卡", "N25"),),
        "max_stamina": (CellSource("简化卡", "N24"),),
        "effect_die": (CellSource("简化卡", "P22"),),
        "auxiliary_die": (CellSource("简化卡", "P23"),),
        "elite_stage": (CellSource("主卡", "AR10"),),
        "level": (CellSource("主卡", "AX10"), CellSource("简化卡", "H7")),
        "profession": (CellSource("主卡", "D9"),),
        "subprofession": (CellSource("主卡", "H9"),),
        "armor_type": (CellSource("职业计算表", "G3"),),
    },
    defaults={
        "speed": 5,
        "elemental_tenacity": 6,
        "weight": 0,
        "current_sp": 0,
        "max_sp": 9,
        "current_stamina": 0,
        "max_stamina": 0,
        "effect_die": "",
        "auxiliary_die": "",
        "elite_stage": 0,
        "armor_type": "轻甲",
    },
)

CARD_PROFILES = {
    RuleMode.V0_3: V03_PROFILE,
    RuleMode.V1_2: V12_PROFILE,
}

ELITE_MAP = {
    "阶级零": 0,
    "阶级一": 1,
    "阶级二": 2,
    "阶段零": 0,
    "阶段一": 1,
    "阶段二": 2,
    "精英零": 0,
    "精英一": 1,
    "精英二": 2,
    "精零": 0,
    "精一": 1,
    "精二": 2,
    "0": 0,
    "1": 1,
    "2": 2,
}


def detect_character_card_rule_mode(filepath: str) -> RuleMode:
    workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=False)
    try:
        sheet_names = set(workbook.sheetnames)
    finally:
        workbook.close()

    matches = [
        profile.rule_mode
        for profile in CARD_PROFILES.values()
        if profile.required_sheets.issubset(sheet_names)
    ]
    if len(matches) != 1:
        raise ValueError("无法根据工作表结构识别角色卡版本，请明确选择 v0.3 或 v1.2")
    return matches[0]


def _first_value(workbook, sources: tuple[CellSource, ...]):
    for source in sources:
        if source.sheet not in workbook.sheetnames:
            continue
        value = workbook[source.sheet][source.cell].value
        if value not in (None, ""):
            return value
    return None


def _extract_name(value) -> str:
    if value is None:
        return "未命名角色"
    text = str(value).strip()
    return text or "未命名角色"


def _extract_number(value, label: str) -> int:
    if value is None:
        raise ValueError(f"{label} 没有可读取的缓存值")
    if isinstance(value, (int, float)):
        return int(round(value))
    text = str(value).strip().lstrip("/")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        raise ValueError(f"{label} 无法解析为数值: {value!r}")
    return int(round(float(match.group(0))))


def _extract_elite(value) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        result = int(value)
        return result if result in (0, 1, 2) else 0
    text = str(value).strip()
    for key, stage in ELITE_MAP.items():
        if key in text:
            return stage
    return 0


def import_character_card_with_report(
    filepath: str,
    rule_mode: RuleMode | str | None = None,
) -> ImportResult:
    detected = detect_character_card_rule_mode(filepath)
    requested = RuleMode.coerce(rule_mode) if rule_mode is not None else detected
    if requested != detected:
        raise ValueError(
            f"所选规则为 v{requested.value}，但文件结构属于 v{detected.value} 角色卡"
        )

    profile = CARD_PROFILES[detected]
    workbook = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
    try:
        raw = {
            field: _first_value(workbook, sources)
            for field, sources in profile.fields.items()
        }
    finally:
        workbook.close()

    values = dict(profile.defaults)
    values.update({key: value for key, value in raw.items() if value not in (None, "")})
    warnings = []

    name = _extract_name(values.get("name"))
    if name == "未命名角色":
        warnings.append("角色名称为空，导入后请手动补充")

    max_hp = _extract_number(values.get("max_hp"), "生命值上限")
    if max_hp <= 0:
        raise ValueError(
            "生命值上限为 0。该角色卡可能尚未填写，或公式缓存未保存；"
            "请在 Excel/WPS 中完成角色并保存后重试。"
        )

    physical_resist = _extract_number(values.get("physical_resist"), "物理抗性")
    magic_resist = _extract_number(values.get("magic_resist"), "法术抗性")
    speed = _extract_number(values.get("speed"), "速度/反应机动")
    tenacity = _extract_number(values.get("elemental_tenacity"), "元素韧性")
    armor_type = str(values.get("armor_type") or profile.defaults["armor_type"]).strip()
    level = _extract_number(values.get("level", 1), "等级")
    profession = str(values.get("profession") or "").strip()
    subprofession = str(values.get("subprofession") or "").strip()

    if detected == RuleMode.V0_3:
        sp = _extract_number(values.get("sp", 0), "SP")
        unit = Unit(
            name=name,
            unit_type="player",
            current_hp=max_hp,
            max_hp=max_hp,
            initial_max_hp=max_hp,
            speed=speed,
            reaction_mobility=_extract_number(
                values.get("reaction_mobility", speed), "反应机动"
            ),
            physical_resist=physical_resist,
            magic_resist=magic_resist,
            armor_type=armor_type,
            elemental_tenacity_current=tenacity,
            elemental_tenacity_max=tenacity,
            current_sp=sp,
            max_sp=sp,
            profession=profession,
            level=level,
        )
    else:
        current_sp = _extract_number(values.get("current_sp", 0), "当前 SP")
        max_sp = _extract_number(values.get("max_sp", 9), "SP 上限")
        elite_stage = _extract_elite(values.get("elite_stage"))
        weight = _extract_number(values.get("weight", 0), "重量等级")
        current_stamina = _extract_number(values.get("current_stamina", 0), "当前耐力")
        max_stamina = _extract_number(values.get("max_stamina", 0), "耐力上限")
        unit = Unit(
            name=name,
            unit_type="player",
            current_hp=max_hp,
            max_hp=max_hp,
            initial_max_hp=max_hp,
            speed=speed,
            reaction_mobility=speed,
            physical_resist=physical_resist,
            magic_resist=magic_resist,
            armor_type=armor_type,
            weight=weight,
            elite_stage=elite_stage,
            elemental_tenacity_current=tenacity,
            elemental_tenacity_max=tenacity,
            current_sp=min(current_sp, max_sp),
            max_sp=max_sp,
            current_stamina=min(current_stamina, max_stamina),
            max_stamina=max_stamina,
            effect_die=str(values.get("effect_die") or "").strip(),
            auxiliary_die=str(values.get("auxiliary_die") or "").strip(),
            profession=profession,
            subprofession=subprofession,
            level=level,
        )

    imported = {
        "name": name,
        "max_hp": max_hp,
        "physical_resist": physical_resist,
        "magic_resist": magic_resist,
        "speed": speed,
        "elemental_tenacity": tenacity,
        "level": values.get("level"),
        "armor_type": armor_type,
        "profession": profession,
        "subprofession": subprofession,
    }
    return ImportResult(unit, detected, imported, tuple(warnings))


def import_character_card(
    filepath: str,
    rule_mode: RuleMode | str | None = None,
) -> Unit:
    """Compatibility wrapper returning only the imported unit."""
    return import_character_card_with_report(filepath, rule_mode).unit


QUICK_IMPORT_FIELDS = [
    ("生命值上限", "max_hp"),
    ("生命值", "max_hp"),
    ("物理抗性", "physical_resist"),
    ("法术抗性", "magic_resist"),
    ("元素韧性", "elemental_tenacity"),
    ("反应机动", "speed"),
    ("速度", "speed"),
    ("重量等级", "weight"),
    ("SP上限", "max_sp"),
    ("技力上限", "max_sp"),
    ("SP", "current_sp"),
    ("技力", "current_sp"),
]


def import_from_quick_text(
    text: str,
    rule_mode: RuleMode | str = RuleMode.V1_2,
    name: str = "",
) -> Unit:
    mode = RuleMode.coerce(rule_mode)
    extracted = {}
    for label, key in QUICK_IMPORT_FIELDS:
        match = re.search(rf"{re.escape(label)}\s*[:：]?\s*(-?\d+)", text)
        if match and key not in extracted:
            extracted[key] = int(match.group(1))

    max_hp = max(1, extracted.get("max_hp", 10))
    default_tenacity = 10 if mode == RuleMode.V0_3 else 6
    current_sp = max(0, extracted.get("current_sp", 0))
    max_sp = max(0, extracted.get("max_sp", current_sp if mode == RuleMode.V0_3 else 9))
    return Unit(
        name=name.strip() or "导入角色",
        unit_type="player",
        current_hp=max_hp,
        max_hp=max_hp,
        initial_max_hp=max_hp,
        physical_resist=extracted.get("physical_resist", 0),
        magic_resist=extracted.get("magic_resist", 0),
        speed=extracted.get("speed", 0 if mode == RuleMode.V0_3 else 5),
        reaction_mobility=extracted.get("speed", 0 if mode == RuleMode.V0_3 else 5),
        weight=extracted.get("weight", 0),
        elemental_tenacity_current=extracted.get("elemental_tenacity", default_tenacity),
        elemental_tenacity_max=extracted.get("elemental_tenacity", default_tenacity),
        current_sp=min(current_sp, max_sp),
        max_sp=max_sp,
        elite_stage=0,
    )

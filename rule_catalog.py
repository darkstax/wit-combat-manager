"""Searchable rule reference built from concise summaries and local workbooks.

The workbook readers intentionally extract small, structured reference records. They
do not copy entire sheets into the application or expose hidden calculation tables.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from threading import RLock
from typing import Iterable, Mapping, Sequence
import os
import re

from models import RuleMode


WORKBOOK_FILENAMES = {
    "v03_card": "角色卡 行于泰拉v0.3 (1).xlsx",
    "v12_professions": "v1.2.战斗职业（非法术部分）.xlsx",
    "v12_arts": "v1.2源石技艺.xlsx",
    "v12_card": "v1.2.角色卡.Plus .xlsx",
}

_SPACE_RE = re.compile(r"\s+")
_FORMULA_RE = re.compile(r"^=(?:DISPIMG|_xlfn\.DISPIMG)\(", re.IGNORECASE)
_INLINE_DISPIMG_RE = re.compile(
    r"^=(?:_xlfn\.)?DISPIMG\([^)]*\)\s*[;；,:：-]*\s*",
    re.IGNORECASE,
)


@dataclass(frozen=True, slots=True)
class RuleEntry:
    """One compact, searchable rule or workbook reference record."""

    version: str
    category: str
    title: str
    body: str
    source: str = ""
    keywords: tuple[str, ...] = field(default_factory=tuple)

    def __post_init__(self):
        object.__setattr__(self, "version", RuleMode.coerce(self.version).value)
        object.__setattr__(self, "category", _clean_text(self.category) or "其他")
        object.__setattr__(self, "title", _clean_text(self.title) or "未命名条目")
        object.__setattr__(self, "body", _clean_multiline(self.body))
        object.__setattr__(self, "source", _clean_text(self.source) or "本地资料")

    @property
    def search_text(self) -> str:
        return " ".join(
            (self.title, self.category, self.version, self.source, self.body, *self.keywords)
        ).casefold()


def scan_directory_for_workbooks(directory: str | Path) -> dict[str, Path]:
    """Find the supported workbooks inside one directory, case-insensitively.

    Only files whose names match the supported workbook filenames are returned;
    files that are not present are simply omitted from the result.
    """

    root = Path(directory).expanduser()
    result: dict[str, Path] = {}
    if not root.is_dir():
        return result
    for key, filename in WORKBOOK_FILENAMES.items():
        wanted = filename.casefold()
        for candidate in root.iterdir():
            if candidate.is_file() and candidate.name.casefold() == wanted:
                result[key] = candidate
                break
    return result


def search_entries(
    entries: Iterable[RuleEntry],
    query: str = "",
    version: RuleMode | str | None = None,
    category: str | None = None,
    limit: int | None = 300,
) -> list[RuleEntry]:
    """Filter and rank entries. All query terms must be present."""

    normalized_version = RuleMode.coerce(version).value if version else None
    normalized_category = _clean_text(category) if category else ""
    terms = [part.casefold() for part in _SPACE_RE.split(query.strip()) if part]
    ranked: list[tuple[int, str, RuleEntry]] = []

    for entry in entries:
        if normalized_version and entry.version != normalized_version:
            continue
        if normalized_category and normalized_category not in {"全部", "全部类别"}:
            if entry.category != normalized_category:
                continue
        haystack = entry.search_text
        if terms and not all(term in haystack for term in terms):
            continue

        title = entry.title.casefold()
        score = 0
        for term in terms:
            if title == term:
                score += 100
            elif title.startswith(term):
                score += 50
            elif term in title:
                score += 25
            elif term in entry.category.casefold():
                score += 10
            elif term in " ".join(entry.keywords).casefold():
                score += 6
            else:
                score += 1
        ranked.append((-score, entry.title.casefold(), entry))

    ranked.sort(key=lambda item: (item[0], item[1], item[2].category))
    result = [item[2] for item in ranked]
    return result if limit is None else result[: max(0, limit)]


class RuleCatalog:
    """Thread-safe lazy catalog used by the nonmodal rule browser."""

    def __init__(
        self,
        workbook_paths: Mapping[str, str | os.PathLike] | Sequence[str | os.PathLike] | None = None,
    ):
        self._external_entries: tuple[RuleEntry, ...] = ()
        self._workbook_paths = _normalize_workbook_paths(workbook_paths)
        self._external_loaded = False
        self._loading = False
        self._errors: list[str] = []
        self._lock = RLock()

    @property
    def external_loaded(self) -> bool:
        with self._lock:
            return self._external_loaded

    @property
    def errors(self) -> tuple[str, ...]:
        with self._lock:
            return tuple(self._errors)

    @property
    def workbook_paths(self) -> dict[str, Path]:
        return dict(self._workbook_paths)

    def entries(self) -> tuple[RuleEntry, ...]:
        """Return all currently loaded (external workbook) entries."""

        with self._lock:
            return self._external_entries

    def categories(self, version: RuleMode | str | None = None) -> list[str]:
        normalized_version = RuleMode.coerce(version).value if version else None
        values = {
            entry.category
            for entry in self.entries()
            if not normalized_version or entry.version == normalized_version
        }
        return sorted(values, key=lambda value: value.casefold())

    def get_profession_names(
        self, version: RuleMode | str | None = None
    ) -> list[str]:
        """Distinct profession names for the version, without loading workbooks.

        "职业" entries contribute their title; "职业技艺" entries contribute their
        first keyword (the owning profession / branch sheet name). Names are
        deduplicated by casefold and returned sorted case-insensitively.
        """

        normalized_version = RuleMode.coerce(version).value if version else None
        by_key: dict[str, str] = {}
        for entry in self.entries():
            if normalized_version and entry.version != normalized_version:
                continue
            if entry.category == "职业":
                if entry.title:
                    by_key.setdefault(entry.title.casefold(), entry.title)
            elif entry.category == "职业技艺":
                if entry.keywords and entry.keywords[0]:
                    by_key.setdefault(entry.keywords[0].casefold(), entry.keywords[0])
        return [by_key[key] for key in sorted(by_key)]

    def search(
        self,
        query: str = "",
        version: RuleMode | str | None = None,
        category: str | None = None,
        limit: int | None = 300,
    ) -> list[RuleEntry]:
        return search_entries(self.entries(), query, version, category, limit)

    def load_external(self) -> tuple[RuleEntry, ...]:
        """Read supported XLSX references once. Safe to call from a worker thread."""

        with self._lock:
            if self._external_loaded:
                return self._external_entries
            if self._loading:
                return self._external_entries
            self._loading = True

        loaded: list[RuleEntry] = []
        errors: list[str] = []
        try:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                errors.append(f"无法载入工作簿支持: {exc}")
                load_workbook = None

            if load_workbook is not None:
                readers = {
                    "v03_card": _read_v03_card,
                    "v12_professions": _read_v12_professions,
                    "v12_arts": _read_v12_arts,
                    "v12_card": _read_v12_quick_reference,
                }
                for key, reader in readers.items():
                    path = self._workbook_paths.get(key)
                    if path is None:
                        continue
                    try:
                        loaded.extend(reader(path, load_workbook))
                    except Exception as exc:  # Workbook corruption must not break the UI.
                        errors.append(f"{path.name}: {exc}")

            deduplicated = _deduplicate(loaded)
            with self._lock:
                self._external_entries = tuple(deduplicated)
                self._errors = errors
                self._external_loaded = True
                return self._external_entries
        finally:
            with self._lock:
                self._loading = False


_shared_catalog: RuleCatalog | None = None
_shared_catalog_lock = RLock()


def get_shared_catalog() -> RuleCatalog:
    """Return the process-wide shared catalog, creating it lazily."""

    global _shared_catalog
    with _shared_catalog_lock:
        if _shared_catalog is None:
            _shared_catalog = RuleCatalog()
        return _shared_catalog


def refresh_shared_catalog(
    workbook_paths: Mapping[str, str | os.PathLike] | Sequence[str | os.PathLike] | None,
) -> RuleCatalog:
    """Rebuild the process-wide shared catalog with the given workbook paths.

    Used after the user changes the rulebook directory setting so that
    ``get_shared_catalog()`` observes the new workbook locations.
    """

    global _shared_catalog
    with _shared_catalog_lock:
        _shared_catalog = RuleCatalog(workbook_paths)
        return _shared_catalog


def _normalize_workbook_paths(
    paths: Mapping[str, str | os.PathLike] | Sequence[str | os.PathLike] | None,
) -> dict[str, Path]:
    if paths is None:
        return {}
    if isinstance(paths, Mapping):
        return {
            key: Path(value).expanduser()
            for key, value in paths.items()
            if key in WORKBOOK_FILENAMES and Path(value).expanduser().is_file()
        }

    by_name = {Path(value).name: Path(value).expanduser() for value in paths}
    return {
        key: by_name[filename]
        for key, filename in WORKBOOK_FILENAMES.items()
        if filename in by_name and by_name[filename].is_file()
    }

def _read_v03_card(path: Path, load_workbook) -> list[RuleEntry]:
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if "源石技艺" not in wb.sheetnames:
            return []
        ws = wb["源石技艺"]
        entries: list[RuleEntry] = []
        # The normalized public-facing arts database occupies AT:BC near the top.
        rows = ws.iter_rows(
            min_row=2, max_row=min(ws.max_row, 221), min_col=46, max_col=55,
            values_only=True,
        )
        for values in rows:
            major = _cell_text(values[0])
            minor = _cell_text(values[1])
            title = _cell_text(values[2])
            if not title:
                continue
            fields = [
                ("大类", major), ("小类", minor),
                ("等级", _cell_text(values[3])),
                ("行动", _cell_text(values[4])),
                ("SP", _cell_text(values[5])),
                ("习得", _cell_text(values[6])),
                ("限制", _cell_text(values[7])),
                ("目标", _cell_text(values[8])),
                ("效果", _cell_text(values[9])),
            ]
            entries.append(RuleEntry(
                "0.3", "源石技艺", title, _format_fields(fields), path.name,
                keywords=tuple(value for value in (major, minor) if value),
            ))
        return entries
    finally:
        wb.close()


def _read_v12_professions(path: Path, load_workbook) -> list[RuleEntry]:
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        entries: list[RuleEntry] = []
        for ws in wb.worksheets:
            if ws.sheet_state != "visible" or ws.title.startswith("WpsReserved"):
                continue
            if ws.title in {"草稿"}:
                continue
            rows = list(ws.iter_rows(
                min_row=1, max_row=min(ws.max_row, 18), min_col=1, max_col=36,
                values_only=True,
            ))
            main_class = _cell_text(_matrix_value(rows, 1, 2))
            overview = [("主职业", main_class)] if main_class else []
            for values in rows:
                label = _cell_text(values[0])
                value = _cell_text(values[1])
                if label and value and not _is_formula(value):
                    overview.append((label, value))
            if overview:
                entries.append(RuleEntry(
                    "1.2", "职业", ws.title, _format_fields(overview), path.name,
                    keywords=tuple(value for value in ("战斗职业", main_class) if value),
                ))
            entries.extend(_read_v12_stage_sheet(
                rows, ws.title, path.name, "职业技艺", (12, 17, 22, 27), 33,
            ))
        return entries
    finally:
        wb.close()


def _read_v12_arts(path: Path, load_workbook) -> list[RuleEntry]:
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        entries: list[RuleEntry] = []
        excluded = {"草稿", "空表格", "法术职业", "v1.2更新记录"}
        for ws in wb.worksheets:
            if ws.sheet_state != "visible" or ws.title in excluded or ws.title.startswith("WpsReserved"):
                continue
            rows = list(ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=31,
                values_only=True,
            ))
            entries.extend(_read_v12_stage_sheet(
                rows, ws.title, path.name, "源石技艺", (2, 7, 12, 17), 23,
            ))
            entries.extend(_read_v12_summons(rows, ws.title, path.name))
        return entries
    finally:
        wb.close()


def _read_v12_stage_sheet(
    rows: Sequence[Sequence[object]],
    sheet_title: str,
    source: str,
    category: str,
    starts: Sequence[int],
    passive_value_col: int,
) -> list[RuleEntry]:
    entries: list[RuleEntry] = []
    stages = ((3, "精英化零"), (7, "精英化一"), (11, "精英化二"), (15, "模组"))
    for base_row, stage in stages:
        if base_row > len(rows):
            continue
        for start_col in starts:
            title = _cell_text(_matrix_value(rows, base_row, start_col))
            if not title or _is_formula(title):
                continue
            fields = [
                ("阶段", stage),
                ("主要行动", _cell_text(_matrix_value(rows, base_row + 1, start_col + 1))),
                ("快速行动", _cell_text(_matrix_value(rows, base_row + 1, start_col + 2))),
                ("范围", _cell_text(_matrix_value(rows, base_row + 1, start_col + 4))),
                ("SP", _cell_text(_matrix_value(rows, base_row + 2, start_col + 1))),
                ("耐力", _cell_text(_matrix_value(rows, base_row + 2, start_col + 2))),
                ("目标", _cell_text(_matrix_value(rows, base_row + 2, start_col + 4))),
                ("效果", _cell_text(_matrix_value(rows, base_row + 3, start_col + 1))),
            ]
            entries.append(RuleEntry(
                "1.2", category, title, _format_fields(fields), source,
                keywords=(sheet_title, stage),
            ))

        passive_title = _cell_text(_matrix_value(rows, base_row + 1, passive_value_col))
        passive_effect = _cell_text(_matrix_value(rows, base_row + 2, passive_value_col))
        if passive_title and not _is_formula(passive_title):
            entries.append(RuleEntry(
                "1.2", "被动技艺", passive_title,
                _format_fields((("职业/流派", sheet_title), ("阶段", stage), ("效果", passive_effect))),
                source, keywords=(sheet_title, stage, category),
            ))
    return entries


def _read_v12_summons(
    rows: Sequence[Sequence[object]], sheet_title: str, source: str,
) -> list[RuleEntry]:
    entries: list[RuleEntry] = []
    row = 19
    while row <= len(rows):
        title = _cell_text(_matrix_value(rows, row, 3))
        tag_label = _cell_text(_matrix_value(rows, row, 4))
        if title and tag_label == "标签":
            values: list[tuple[str, str]] = [("流派", sheet_title)]
            tag = _cell_text(_matrix_value(rows, row, 5))
            if tag:
                values.append(("标签", tag))
            next_row = row + 1
            while next_row <= len(rows) and next_row < row + 18:
                if _cell_text(_matrix_value(rows, next_row, 3)) and _cell_text(_matrix_value(rows, next_row, 4)) == "标签":
                    break
                detail = _cell_text(_matrix_value(rows, next_row, 4))
                if detail and not _is_formula(detail):
                    values.append(("属性", detail))
                next_row += 1
            entries.append(RuleEntry(
                "1.2", "召唤物", title, _format_fields(values), source,
                keywords=(sheet_title, tag),
            ))
            row = next_row
            continue
        row += 1
    return entries


def _read_v12_quick_reference(path: Path, load_workbook) -> list[RuleEntry]:
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if "战斗信息速查表" not in wb.sheetnames:
            return []
        ws = wb["战斗信息速查表"]
        entries: list[RuleEntry] = []
        rows = ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=1, max_col=8,
            values_only=True,
        )
        for values in rows:
            action = _cell_text(values[0])
            if action:
                entries.append(RuleEntry(
                    "1.2", "战斗动作", action,
                    _format_fields((
                        ("消耗", _cell_text(values[1])),
                        ("条件", _cell_text(values[2])),
                        ("效果", _cell_text(values[3])),
                    )), path.name,
                ))
            positive = _cell_text(values[4])
            if positive:
                entries.append(RuleEntry(
                    "1.2", "正面状态", positive, _cell_text(values[5]), path.name,
                    keywords=("状态", "增益"),
                ))
            negative = _cell_text(values[6])
            if negative:
                entries.append(RuleEntry(
                    "1.2", "负面状态", negative, _cell_text(values[7]), path.name,
                    keywords=("状态", "减益"),
                ))
        return entries
    finally:
        wb.close()


def _deduplicate(entries: Iterable[RuleEntry]) -> list[RuleEntry]:
    unique: dict[tuple[str, str, str, str], RuleEntry] = {}
    for entry in entries:
        key = (entry.version, entry.category, entry.title.casefold(), entry.body.casefold())
        unique.setdefault(key, entry)
    return list(unique.values())


def _matrix_value(rows: Sequence[Sequence[object]], row: int, column: int) -> object:
    if row < 1 or column < 1 or row > len(rows):
        return None
    values = rows[row - 1]
    return values[column - 1] if column <= len(values) else None


def _format_fields(fields: Iterable[tuple[str, object]]) -> str:
    lines = []
    for label, raw_value in fields:
        value = _cell_text(raw_value)
        if not value or _is_formula(value) or value == "/":
            continue
        lines.append(f"{_clean_text(label)}：{value}")
    return "\n".join(lines)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = _INLINE_DISPIMG_RE.sub("", str(value).strip())
    return _clean_multiline(text)


def _is_formula(value: str) -> bool:
    return bool(_FORMULA_RE.match(value)) or value.startswith("=")


def _clean_text(value: object) -> str:
    return _SPACE_RE.sub(" ", str(value or "")).strip()


def _clean_multiline(value: object) -> str:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
    lines = [_SPACE_RE.sub(" ", line).strip() for line in text.split("\n")]
    return "\n".join(line for line in lines if line)
